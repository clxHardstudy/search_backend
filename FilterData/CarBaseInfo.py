from openpyxl import load_workbook
import json
from datetime import datetime

# 加载 .xlsm 文件并保留 VBA 宏
workbook = load_workbook('../底盘关键性能竞品数据（KC）-20240724.xlsm', keep_vba=True)

# 选择工作表“轿车前”
# sheet = workbook['轿车前']
sheet = workbook['SUV前']

# 获取第四列（D列）开始的前三行数据
data = []
dic = {}

for col in range(4, sheet.max_column + 1):
    for row in range(1, 4):
        if row == 1 and sheet.cell(row=row, column=col).value is None:
            break
        cell_value = sheet.cell(row=row, column=col).value
        if row == 1:
            dic.update({"name": cell_value.replace('\n', '')})
        if row == 2:
            if isinstance(cell_value, str):
                dic.update({"release_date": cell_value})
            if isinstance(cell_value, datetime):
                dic.update({"release_date": cell_value.strftime('%Y-%m-%d')})
            else:
                dic.update({"release_date": None})
        if row == 3:
            dic.update({"wheelbase": cell_value})
    if dic:
        # 轿车：0 、SUV：1
        dic.update(({"car_type_id": 2}))
        data.append(dic)
    dic = {}

data = json.dumps(data, ensure_ascii=False, indent=4)
print(data)
