from openpyxl import load_workbook
import json

# 加载 .xlsm 文件并保留 VBA 宏
workbook = load_workbook('../底盘关键性能竞品数据（KC）-20240724.xlsm', keep_vba=True)

# 创建工作表字典
sheet_dic = {
    "car_front": workbook['轿车前'],
    "car_rear": workbook['轿车后'],
    # "suv_front": workbook['SUV前'],
    # "suv_rear": workbook['SUV后'],
}

# 指标与行号的映射
row_to_metric = {
    5: "wheel_rate",
    6: "vertical_force",
    7: "ride_frequency_no_tyre",
    8: "ride_rate",
    9: "ride_frequency_with_tyre",
    10: "tyre_radial_rate",
    11: "toe_angle_change",
    12: "camber_change",
    13: "spin_angle_change",
    14: "kinematic_roll_centre_height",
    15: "fore_aft_displacement_wc",
    16: "fore_aft_displacement_tcp",
}

# 存储结果的字典
dic = {}
data = {}
# 迭代每个工作表
for col in range(4, 32):
    dic = {}
    for sheet_name, sheet in sheet_dic.items():
        # 这里限制列的范围，如果需要处理多列可以调整范围
        # 迭代行，根据映射字典进行操作
        for row, metric in row_to_metric.items():
            cell_value = sheet.cell(row=row, column=col).value
            # 将当前表名和单元格值作为字典
            cell_value = {sheet_name: cell_value}
            # 插入数据到主字典中
            if metric in dic:
                dic[metric].update(cell_value)
            else:
                dic[metric] = cell_value
    data.update({str(col - 3): dic})

# 输出结果
final_data = {}
final_data["verticalParallel_ARBConnected"] = data
data = json.dumps(final_data, ensure_ascii=False, indent=4)
print(data)
